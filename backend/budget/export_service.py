"""
BudgetFlow — Service d'export PDF et Excel pour les rapports budgétaires.
"""
import io
from decimal import Decimal
from datetime import datetime

from django.http import HttpResponse


def _fmt(n):
    """Formate un nombre en FCFA."""
    try:
        return f"{float(n or 0):,.0f} FCFA".replace(',', ' ')
    except (TypeError, ValueError):
        return '0 FCFA'


def _taux_bar(taux, width=20):
    """Barre ASCII pour représenter un taux d'utilisation."""
    filled = int(min(float(taux or 0), 100) / 100 * width)
    return '█' * filled + '░' * (width - filled)


# ─── Export Excel ──────────────────────────────────────────────────────────────

def _export_excel(rapport):
    try:
        import openpyxl
        from openpyxl.styles import (
            Font, PatternFill, Alignment, Border, Side, numbers
        )
        from openpyxl.utils import get_column_letter
    except ImportError:
        raise ImportError("openpyxl est requis pour l'export Excel. Installez-le avec : pip install openpyxl")

    wb = openpyxl.Workbook()
    meta   = rapport.get('meta', {})
    resume = rapport.get('resume', {})

    # ── Couleurs ──────────────────────────────────────────────────────────────
    NAVY   = '0D1E35'
    GOLD   = 'C9A84C'
    LIGHT  = 'F8F9FA'
    WHITE  = 'FFFFFF'
    RED    = 'DC2626'
    ORANGE = 'D97706'
    GREEN  = '059669'

    def _header_style(cell, bg=NAVY, fg=WHITE, bold=True, size=11):
        cell.font      = Font(bold=bold, color=fg, size=size, name='Calibri')
        cell.fill      = PatternFill('solid', fgColor=bg)
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

    def _border():
        s = Side(style='thin', color='D1D5DB')
        return Border(left=s, right=s, top=s, bottom=s)

    def _money(cell, value):
        try:
            cell.value = float(value or 0)
        except (TypeError, ValueError):
            cell.value = 0
        cell.number_format = '#,##0" FCFA"'
        cell.alignment = Alignment(horizontal='right')

    def _pct(cell, value):
        try:
            cell.value = float(value or 0) / 100
        except (TypeError, ValueError):
            cell.value = 0
        cell.number_format = '0.00%'
        cell.alignment = Alignment(horizontal='center')

    # ── Feuille 1 : Résumé ────────────────────────────────────────────────────
    ws = wb.active
    ws.title = 'Résumé'

    # Titre
    ws.merge_cells('A1:F1')
    title_cell = ws['A1']
    title_cell.value = f"BudgetFlow — Rapport {meta.get('type', '')} : {meta.get('label_periode', '')}"
    _header_style(title_cell, NAVY, GOLD, bold=True, size=14)
    ws.row_dimensions[1].height = 32

    ws.merge_cells('A2:F2')
    sub_cell = ws['A2']
    sub_cell.value = f"Période : {meta.get('date_debut', '')} → {meta.get('date_fin', '')}   |   Généré le : {meta.get('genere_le', '')[:19]}"
    sub_cell.font      = Font(color='6B7280', size=10, italic=True, name='Calibri')
    sub_cell.alignment = Alignment(horizontal='center')
    ws.row_dimensions[2].height = 20

    # KPIs
    ws.row_dimensions[4].height = 22
    headers = ['Indicateur', 'Valeur']
    for col, h in enumerate(headers, 1):
        c = ws.cell(row=4, column=col, value=h)
        _header_style(c, GOLD, NAVY)

    kpis = [
        ('Budgets actifs',        resume.get('nb_budgets', 0)),
        ('Montant global',        _fmt(resume.get('montant_global'))),
        ('Montant consommé',      _fmt(resume.get('montant_consomme'))),
        ('Montant disponible',    _fmt(resume.get('montant_disponible'))),
        ('Taux de consommation',  f"{resume.get('taux_global', 0):.2f} %"),
        ('Dépenses sur période',  rapport.get('nb_depenses_periode', 0)),
        ('Total dépenses',        _fmt(rapport.get('total_depenses_periode'))),
    ]
    nb_par_statut = resume.get('nb_par_statut', {})
    for s_key, s_label in [('APPROUVE', 'Budgets approuvés'), ('SOUMIS', 'En attente'),
                            ('REJETE', 'Rejetés'), ('CLOTURE', 'Clôturés')]:
        kpis.append((s_label, nb_par_statut.get(s_key, 0)))

    for row_i, (label, val) in enumerate(kpis, 5):
        ws.cell(row=row_i, column=1, value=label).font = Font(name='Calibri', size=10)
        ws.cell(row=row_i, column=2, value=val).font   = Font(name='Calibri', size=10, bold=True)

    ws.column_dimensions['A'].width = 28
    ws.column_dimensions['B'].width = 22

    # ── Feuille 2 : Dépenses par département ──────────────────────────────────
    ws2 = wb.create_sheet('Par département')
    ws2.merge_cells('A1:D1')
    t2 = ws2['A1']
    t2.value = 'Dépenses par département'
    _header_style(t2, NAVY, GOLD, size=12)
    ws2.row_dimensions[1].height = 26

    cols2 = ['Département', 'Total consommé (FCFA)', 'Nb dépenses', 'Part (%)']
    for col, h in enumerate(cols2, 1):
        c = ws2.cell(row=2, column=col, value=h)
        _header_style(c, GOLD, NAVY)

    dept_data   = rapport.get('depenses_par_departement', [])
    total_global = float(rapport.get('total_depenses_periode') or 1)
    for ri, d in enumerate(dept_data, 3):
        nom   = d.get('ligne__budget__departement__nom') or '—'
        total = float(d.get('total') or 0)
        nb    = d.get('nb', 0)
        part  = total / total_global * 100 if total_global else 0
        ws2.cell(row=ri, column=1, value=nom).font = Font(name='Calibri', size=10)
        c2 = ws2.cell(row=ri, column=2)
        _money(c2, total)
        ws2.cell(row=ri, column=3, value=nb).alignment = Alignment(horizontal='center')
        c4 = ws2.cell(row=ri, column=4)
        _pct(c4, part)

    for col_i, width in enumerate([30, 22, 14, 12], 1):
        ws2.column_dimensions[get_column_letter(col_i)].width = width

    # ── Feuille 3 : Top dépenses ──────────────────────────────────────────────
    ws3 = wb.create_sheet('Top dépenses')
    ws3.merge_cells('A1:E1')
    t3 = ws3['A1']
    t3.value = 'Top lignes budgétaires consommatrices'
    _header_style(t3, NAVY, GOLD, size=12)

    cols3 = ['Code budget', 'Nom budget', 'Ligne budgétaire', 'Total (FCFA)', 'Nb']
    for col, h in enumerate(cols3, 1):
        _header_style(ws3.cell(row=2, column=col, value=h), GOLD, NAVY)

    for ri, d in enumerate(rapport.get('top_depenses', []), 3):
        ws3.cell(row=ri, column=1, value=d.get('ligne__budget__code', ''))
        ws3.cell(row=ri, column=2, value=d.get('ligne__budget__nom', ''))
        ws3.cell(row=ri, column=3, value=d.get('ligne__libelle', ''))
        _money(ws3.cell(row=ri, column=4), d.get('total'))
        ws3.cell(row=ri, column=5, value=d.get('nb', 0)).alignment = Alignment(horizontal='center')

    for col_i, width in enumerate([14, 28, 36, 20, 8], 1):
        ws3.column_dimensions[get_column_letter(col_i)].width = width

    # ── Feuille 4 : Alertes ───────────────────────────────────────────────────
    alertes = rapport.get('alertes', [])
    if alertes:
        ws4 = wb.create_sheet('Alertes')
        ws4.merge_cells('A1:F1')
        t4 = ws4['A1']
        t4.value = f'Alertes budgétaires ({len(alertes)} budget(s) en dépassement)'
        _header_style(t4, RED, WHITE, size=12)

        cols4 = ['Code', 'Nom', 'Département', 'Taux %', 'Consommé', 'Niveau']
        for col, h in enumerate(cols4, 1):
            _header_style(ws4.cell(row=2, column=col, value=h), NAVY, WHITE)

        NIVEAU_COLORS = {'CRITIQUE': 'DC2626', 'ROUGE': 'EF4444', 'ORANGE': 'F59E0B'}
        for ri, a in enumerate(alertes, 3):
            ws4.cell(row=ri, column=1, value=a['budget_code'])
            ws4.cell(row=ri, column=2, value=a['budget_nom'])
            ws4.cell(row=ri, column=3, value=a['departement'])
            _pct(ws4.cell(row=ri, column=4), a['taux'])
            _money(ws4.cell(row=ri, column=5), a['montant_consomme'])
            nc = ws4.cell(row=ri, column=6, value=a['niveau'])
            nc.font = Font(bold=True, color=NIVEAU_COLORS.get(a['niveau'], RED), name='Calibri')

        for col_i, width in enumerate([14, 32, 22, 10, 22, 12], 1):
            ws4.column_dimensions[get_column_letter(col_i)].width = width

    # ── Sérialisation ──────────────────────────────────────────────────────────
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    label   = meta.get('label_periode', 'rapport').replace(' ', '_').replace('/', '-')
    filename = f"rapport_{meta.get('type', 'adhoc').lower()}_{label}.xlsx"

    response = HttpResponse(
        buf.getvalue(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    return response


# ─── Export PDF ────────────────────────────────────────────────────────────────

def _export_pdf(rapport):
    try:
        from reportlab.lib.pagesizes import A4, landscape
        from reportlab.lib import colors
        from reportlab.lib.units import cm, mm
        from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
        from reportlab.platypus import (
            SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle,
            HRFlowable, KeepTogether,
        )
        from reportlab.lib.enums import TA_CENTER, TA_LEFT, TA_RIGHT
    except ImportError:
        raise ImportError("reportlab est requis pour l'export PDF. Installez-le avec : pip install reportlab")

    meta   = rapport.get('meta', {})
    resume = rapport.get('resume', {})

    buf = io.BytesIO()
    doc = SimpleDocTemplate(
        buf,
        pagesize=A4,
        leftMargin=1.8 * cm, rightMargin=1.8 * cm,
        topMargin=2 * cm,    bottomMargin=2 * cm,
    )

    # ── Palette ───────────────────────────────────────────────────────────────
    NAVY   = colors.HexColor('#0D1E35')
    GOLD   = colors.HexColor('#C9A84C')
    LIGHT  = colors.HexColor('#F9FAFB')
    RED    = colors.HexColor('#DC2626')
    ORANGE = colors.HexColor('#D97706')
    GREEN  = colors.HexColor('#059669')
    GRAY   = colors.HexColor('#6B7280')
    WHITE  = colors.white

    styles = getSampleStyleSheet()
    style_title  = ParagraphStyle('Title',  fontSize=18, textColor=WHITE,  alignment=TA_CENTER, fontName='Helvetica-Bold', leading=22)
    style_sub    = ParagraphStyle('Sub',    fontSize=10, textColor=colors.HexColor('#C9A84C'), alignment=TA_CENTER, fontName='Helvetica', leading=14)
    style_h2     = ParagraphStyle('H2',     fontSize=13, textColor=NAVY, fontName='Helvetica-Bold', leading=18, spaceBefore=14)
    style_body   = ParagraphStyle('Body',   fontSize=9,  textColor=colors.HexColor('#374151'), fontName='Helvetica', leading=13)
    style_right  = ParagraphStyle('Right',  fontSize=9,  textColor=colors.HexColor('#374151'), fontName='Helvetica', alignment=TA_RIGHT)

    story = []

    # ── En-tête ───────────────────────────────────────────────────────────────
    header_data = [[Paragraph(
        f"BudgetFlow — Rapport {meta.get('type', '')} : {meta.get('label_periode', '')}",
        style_title
    )]]
    header_table = Table(header_data, colWidths=[doc.width])
    header_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, -1), NAVY),
        ('TOPPADDING',  (0, 0), (-1, -1), 14),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 14),
        ('ROUNDEDCORNERS', [8]),
    ]))
    story.append(header_table)
    story.append(Spacer(1, 6))

    story.append(Paragraph(
        f"Période : {meta.get('date_debut', '')} → {meta.get('date_fin', '')}   |   "
        f"Généré le : {meta.get('genere_le', '')[:19]}",
        style_sub,
    ))
    story.append(Spacer(1, 16))

    # ── KPI Cards ─────────────────────────────────────────────────────────────
    nb_par_statut = resume.get('nb_par_statut', {})
    kpi_data = [
        ['Budgets',                      str(resume.get('nb_budgets', 0))],
        ['Dépenses (période)',            str(rapport.get('nb_depenses_periode', 0))],
        ['Montant global',                _fmt(resume.get('montant_global'))],
        ['Montant consommé',              _fmt(resume.get('montant_consomme'))],
        ['Montant disponible',            _fmt(resume.get('montant_disponible'))],
        ['Taux consommation',             f"{resume.get('taux_global', 0):.2f} %"],
        ['Total dépenses période',        _fmt(rapport.get('total_depenses_periode'))],
        ['Budgets approuvés',             str(nb_par_statut.get('APPROUVE', 0))],
        ['Budgets en attente',            str(nb_par_statut.get('SOUMIS', 0))],
        ['Budgets rejetés',               str(nb_par_statut.get('REJETE', 0))],
    ]

    kpi_rows = []
    for i in range(0, len(kpi_data), 2):
        row = kpi_data[i:i+2]
        while len(row) < 2:
            row.append(['', ''])
        kpi_rows.append([
            Paragraph(f"<b>{row[0][0]}</b>", style_body),
            Paragraph(row[0][1], ParagraphStyle('KV', fontSize=10, fontName='Helvetica-Bold', textColor=NAVY, alignment=TA_RIGHT)),
            Paragraph(f"<b>{row[1][0]}</b>", style_body),
            Paragraph(row[1][1], ParagraphStyle('KV', fontSize=10, fontName='Helvetica-Bold', textColor=NAVY, alignment=TA_RIGHT)),
        ])

    col_w = doc.width / 4
    kpi_table = Table(kpi_rows, colWidths=[col_w * 1.4, col_w * 0.6, col_w * 1.4, col_w * 0.6])
    kpi_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, -1), LIGHT),
        ('ROWBACKGROUNDS', (0, 0), (-1, -1), [WHITE, LIGHT]),
        ('BOX',     (0, 0), (-1, -1), 0.5, colors.HexColor('#E5E7EB')),
        ('INNERGRID', (0, 0), (-1, -1), 0.3, colors.HexColor('#F3F4F6')),
        ('TOPPADDING',    (0, 0), (-1, -1), 7),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 7),
        ('LEFTPADDING',   (0, 0), (-1, -1), 10),
        ('RIGHTPADDING',  (0, 0), (-1, -1), 10),
    ]))
    story.append(kpi_table)
    story.append(Spacer(1, 18))

    # ── Dépenses par département ───────────────────────────────────────────────
    dept_data = rapport.get('depenses_par_departement', [])
    if dept_data:
        story.append(Paragraph('Dépenses par département', style_h2))
        story.append(HRFlowable(width='100%', thickness=1, color=GOLD, spaceAfter=8))

        total_global = float(rapport.get('total_depenses_periode') or 1)
        dept_rows = [[
            Paragraph('<b>Département</b>', style_body),
            Paragraph('<b>Total consommé</b>', style_right),
            Paragraph('<b>Nb dépenses</b>', style_right),
            Paragraph('<b>Part</b>', style_right),
        ]]
        for d in dept_data:
            nom   = d.get('ligne__budget__departement__nom') or '—'
            total = float(d.get('total') or 0)
            nb    = d.get('nb', 0)
            part  = total / total_global * 100 if total_global else 0
            dept_rows.append([
                Paragraph(nom, style_body),
                Paragraph(_fmt(total), style_right),
                Paragraph(str(nb), style_right),
                Paragraph(f'{part:.1f} %', style_right),
            ])

        dept_table = Table(dept_rows, colWidths=[doc.width * .35, doc.width * .30, doc.width * .18, doc.width * .17])
        dept_table.setStyle(TableStyle([
            ('BACKGROUND',    (0, 0), (-1, 0), NAVY),
            ('TEXTCOLOR',     (0, 0), (-1, 0), WHITE),
            ('ROWBACKGROUNDS', (0, 1), (-1, -1), [WHITE, LIGHT]),
            ('BOX',           (0, 0), (-1, -1), 0.5, colors.HexColor('#E5E7EB')),
            ('INNERGRID',     (0, 0), (-1, -1), 0.3, colors.HexColor('#F3F4F6')),
            ('TOPPADDING',    (0, 0), (-1, -1), 6),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 6),
            ('LEFTPADDING',   (0, 0), (-1, -1), 10),
        ]))
        story.append(dept_table)
        story.append(Spacer(1, 18))

    # ── Top dépenses ──────────────────────────────────────────────────────────
    top_data = rapport.get('top_depenses', [])
    if top_data:
        story.append(Paragraph('Top lignes budgétaires consommatrices', style_h2))
        story.append(HRFlowable(width='100%', thickness=1, color=GOLD, spaceAfter=8))

        top_rows = [[
            Paragraph('<b>#</b>', style_body),
            Paragraph('<b>Budget</b>', style_body),
            Paragraph('<b>Ligne</b>', style_body),
            Paragraph('<b>Consommé</b>', style_right),
            Paragraph('<b>Nb</b>', style_right),
        ]]
        for idx, d in enumerate(top_data, 1):
            top_rows.append([
                Paragraph(str(idx), style_body),
                Paragraph(f"{d.get('ligne__budget__code', '')} – {d.get('ligne__budget__nom', '')}", style_body),
                Paragraph(d.get('ligne__libelle', ''), style_body),
                Paragraph(_fmt(d.get('total')), style_right),
                Paragraph(str(d.get('nb', 0)), style_right),
            ])

        top_table = Table(top_rows, colWidths=[
            doc.width * .04, doc.width * .30, doc.width * .38, doc.width * .20, doc.width * .08,
        ])
        top_table.setStyle(TableStyle([
            ('BACKGROUND',    (0, 0), (-1, 0), NAVY),
            ('TEXTCOLOR',     (0, 0), (-1, 0), WHITE),
            ('ROWBACKGROUNDS', (0, 1), (-1, -1), [WHITE, LIGHT]),
            ('BOX',           (0, 0), (-1, -1), 0.5, colors.HexColor('#E5E7EB')),
            ('INNERGRID',     (0, 0), (-1, -1), 0.3, colors.HexColor('#F3F4F6')),
            ('TOPPADDING',    (0, 0), (-1, -1), 5),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 5),
            ('LEFTPADDING',   (0, 0), (-1, -1), 8),
        ]))
        story.append(top_table)
        story.append(Spacer(1, 18))

    # ── Alertes ────────────────────────────────────────────────────────────────
    alertes = rapport.get('alertes', [])
    if alertes:
        story.append(Paragraph(f'Alertes budgétaires ({len(alertes)} budget(s))', style_h2))
        story.append(HRFlowable(width='100%', thickness=1, color=RED, spaceAfter=8))

        NIVEAU_COLORS_PDF = {
            'CRITIQUE': colors.HexColor('#DC2626'),
            'ROUGE':    colors.HexColor('#EF4444'),
            'ORANGE':   colors.HexColor('#F59E0B'),
        }

        alerte_rows = [[
            Paragraph('<b>Code</b>', style_body),
            Paragraph('<b>Nom</b>', style_body),
            Paragraph('<b>Département</b>', style_body),
            Paragraph('<b>Taux</b>', style_right),
            Paragraph('<b>Consommé</b>', style_right),
            Paragraph('<b>Niveau</b>', style_right),
        ]]
        for a in alertes:
            nc = NIVEAU_COLORS_PDF.get(a['niveau'], RED)
            alerte_rows.append([
                Paragraph(a['budget_code'], style_body),
                Paragraph(a['budget_nom'], style_body),
                Paragraph(a['departement'], style_body),
                Paragraph(f"{a['taux']:.1f} %", style_right),
                Paragraph(_fmt(a['montant_consomme']), style_right),
                Paragraph(f'<font color="#{a["niveau"] and nc.hexval()[2:]}"><b>{a["niveau"]}</b></font>', style_right),
            ])

        alerte_table = Table(alerte_rows, colWidths=[
            doc.width * .10, doc.width * .27, doc.width * .22,
            doc.width * .09, doc.width * .20, doc.width * .12,
        ])
        alerte_table.setStyle(TableStyle([
            ('BACKGROUND',    (0, 0), (-1, 0), colors.HexColor('#DC2626')),
            ('TEXTCOLOR',     (0, 0), (-1, 0), WHITE),
            ('ROWBACKGROUNDS', (0, 1), (-1, -1), [WHITE, colors.HexColor('#FFF1F2')]),
            ('BOX',           (0, 0), (-1, -1), 0.5, colors.HexColor('#FECDD3')),
            ('INNERGRID',     (0, 0), (-1, -1), 0.3, colors.HexColor('#FEE2E2')),
            ('TOPPADDING',    (0, 0), (-1, -1), 5),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 5),
            ('LEFTPADDING',   (0, 0), (-1, -1), 8),
        ]))
        story.append(alerte_table)

    # ── Pied de page ──────────────────────────────────────────────────────────
    story.append(Spacer(1, 20))
    story.append(HRFlowable(width='100%', thickness=0.5, color=colors.HexColor('#E5E7EB'), spaceAfter=6))
    story.append(Paragraph(
        f"BudgetFlow — Système de gestion budgétaire   |   Document généré automatiquement",
        ParagraphStyle('Footer', fontSize=8, textColor=GRAY, alignment=TA_CENTER, fontName='Helvetica-Oblique'),
    ))

    doc.build(story)
    buf.seek(0)

    label    = meta.get('label_periode', 'rapport').replace(' ', '_').replace('/', '-')
    filename = f"rapport_{meta.get('type', 'adhoc').lower()}_{label}.pdf"

    response = HttpResponse(buf.getvalue(), content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    return response


# ─── Facade ────────────────────────────────────────────────────────────────────

class ExportService:
    @staticmethod
    def export_excel(rapport):
        return _export_excel(rapport)

    @staticmethod
    def export_pdf(rapport):
        return _export_pdf(rapport)
